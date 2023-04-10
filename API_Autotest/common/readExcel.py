"""
读写Excel
"""
import os
import openpyxl
from openpyxl.styles import PatternFill



class RwExcel():
    def __init__(self,fileName,sheetName):

        self.fileName = fileName
        self.sheetName = sheetName

    def open(self):
        self.workbook = openpyxl.load_workbook(self.fileName)
        self.sheetname = self.workbook[self.sheetName]

    def read(self):
        """
        读取Excel数据
        :return:列表嵌套字典
        """
        self.open()
        max_row = self.sheetname.max_row
        max_column = self.sheetname.max_column
        datas=[]
        for row in range(1,max_row+1):
            data_list = []
            for column in range(1,max_column+1):
                data=self.sheetname.cell(row=row,column=column).value
                data_list.append(data)
            datas.append(data_list)
        title = datas[0]
        cases = []
        for element in datas[1:]:
            new_data=dict(zip(title,element))
            cases.append(new_data)
        return cases

    def write(self,row,value,column=8):
        """
        :param row: 行，调用时可根据模板中case_id  模板：row=case_id+1
        :param value: 写入的内容
        :param column: 列，模板固定第8列
        :return:
        """
        self.open()
        # Color=['c6efce','006100']#绿
        # Color = ['ffc7ce', '9c0006']  #红
        Colors=[
            'c6efce',
            'ff3300'
        ]
        if value == "pass":
            fille = PatternFill('solid', fgColor=Colors[0])
        else:
            fille = PatternFill('solid', fgColor=Colors[1])
        self.sheetname.cell(row=row,column=column,value=value).fill=fille
        self.sava()

    def sava(self):
        """

        :return:
        """
        self.workbook.save(self.fileName)




# if __name__ == '__main__':
#     path=os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
#     path1=os.path.join(path,"datas")
#     fileName = os.path.join(path1,"IMS.xlsx")
#     res=RwExcel(fileName,"Sheet1")
#     print(res.read())

